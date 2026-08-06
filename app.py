import streamlit as st
import pandas as pd
import anthropic
import json
import os
import re
from datetime import datetime

# ─────────────────────────────────────────
# 사업장별 키워드 매핑
# ─────────────────────────────────────────
SITE_KEYWORDS = {
    "SR6": {"salary": ["삼성전자기흥", "SR6"], "severance": ["산업체"]},
    "동탄에듀": {"salary": ["동탄", "에듀"], "severance": ["연수원"]},
    "넥슨": {"salary": ["넥슨"], "severance": ["오피스"]}
}

def extract_month_from_filename(filename):
    """파일명에서 월 추출 (예: 2026.04 → 2026년 04월)"""
    match = re.search(r'(\d{4})[.\-_](\d{2})', str(filename))
    if match:
        return f"{match.group(1)}년 {match.group(2)}월"
    return None

def filter_site_data(df, site_name):
    if df is None or df.empty:
        return pd.DataFrame(), pd.DataFrame()
    keywords = SITE_KEYWORDS.get(site_name, {})
    site_col = df.columns[1]
    salary_df = df[df[site_col].apply(lambda x: any(kw in str(x) for kw in keywords.get("salary", [])) if pd.notna(x) else False)]
    sev_df = df[df[site_col].apply(lambda x: any(kw in str(x) for kw in keywords.get("severance", [])) if pd.notna(x) else False)]
    return salary_df, sev_df

def read_excel_sheets(file):
    import io, os
    os.environ["PYTHONIOENCODING"] = "utf-8"
    engine = "xlrd" if str(file.name).endswith(".xls") else "openpyxl"
    file_bytes = file.read()
    xl = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
    sheets = {}
    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, engine=engine, header=4)
            df.columns = [str(c).encode("utf-8", errors="replace").decode("utf-8") for c in df.columns]
            # 모든 문자열 컬럼 utf-8 처리
            for col in df.select_dtypes(include="object").columns:
                df[col] = df[col].apply(lambda x: str(x).encode("utf-8", errors="replace").decode("utf-8") if pd.notna(x) else x)
            sheets[sheet] = df
        except Exception as e:
            pass
    file.seek(0)
    return sheets

def read_extra_cost_sheet(file, site_name):
    try:
        import io
        engine = "xlrd" if str(file.name).endswith(".xls") else "openpyxl"
        file_bytes = file.read()
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="기타비용", engine=engine, header=0)
        file.seek(0)
        df.columns = [str(c) for c in df.columns]
        site_col = df.columns[1]
        keywords = SITE_KEYWORDS.get(site_name, {}).get("salary", [])
        filtered = df[df[site_col].apply(lambda x: any(kw in str(x) for kw in keywords) if pd.notna(x) else False)]
        if filtered.empty:
            return "해당 사업장 기타비용 없음"
        result = []
        for _, row in filtered.iterrows():
            cols = list(row.index)
            name_col = cols[3] if len(cols) > 3 else None
            amount_col = cols[12] if len(cols) > 12 else None
            reason_col = cols[13] if len(cols) > 13 else None
            name = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else ""
            amount = row[amount_col] if amount_col and pd.notna(row[amount_col]) else 0
            reason = str(row[reason_col]).strip() if reason_col and pd.notna(row[reason_col]) else ""
            site = str(row[site_col]).strip() if pd.notna(row[site_col]) else ""
            if name and name not in ["nan", ""]:
                result.append(f"  - [{site}] {name} / 사유: {reason} / 금액: {int(float(amount)):,}원")
            else:
                result.append(f"  - [{site}] 보험료정산 / 금액: {int(float(amount)):,}원")
        return "\n".join(result) if result else "기타비용 없음"
    except Exception as e:
        return f"기타비용 시트 읽기 오류: {str(e)}"

def read_profit_file(file):
    try:
        df = pd.read_excel(file)
        month_cols = [c for c in df.columns if '년' in str(c) and '월' in str(c)]
        if not month_cols:
            return df.to_string(), None, None
        non_zero_months = []
        for c in month_cols:
            try:
                col_sum = pd.to_numeric(df[c], errors='coerce').sum()
                if col_sum != 0:
                    non_zero_months.append(c)
            except:
                pass
        if len(non_zero_months) < 2:
            return df.to_string(), None, None
        prev_month = non_zero_months[-2]
        curr_month = non_zero_months[-1]
        id_cols = [c for c in df.columns if '년' not in str(c) and '월' not in str(c) and 'Unnamed' not in str(c)]
        selected = df[id_cols + [prev_month, curr_month]].copy()
        selected['증감'] = pd.to_numeric(selected[curr_month], errors='coerce') - pd.to_numeric(selected[prev_month], errors='coerce')
        result = f"[전월: {prev_month} / 당월: {curr_month}]\n\n" + selected.to_string()
        return result, prev_month, curr_month
    except Exception as e:
        return f"읽기 오류: {str(e)}", None, None

# ─────────────────────────────────────────
# 급여내역서 생성 (총인건비 파일 → 사업장별 급여내역서)
# ─────────────────────────────────────────
PAYSLIP_HEADERS = [
    "사업장", "성명", "기본급", "고정\n연장수당", "고정\n야간수당", "직책\n수당",
    "기타\n수당", "연차\n수당", "주휴\n수당", "야간\n수당", "휴일\n수당",
    "연장\n수당", "가지급", "퇴사자연차수당", "인센티브", "지급\n액계",
]
PAYSLIP_FONT = "맑은 고딕"


def extract_ym_from_filename(filename):
    """파일명에서 연/월 추출 (예: 2026_06 → (2026, 6))"""
    m = re.search(r"(20\d{2})[._\-]?(0[1-9]|1[0-2])", str(filename))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def read_total_labor_file(file_bytes):
    """총인건비 파일 로드"""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "급여" not in wb.sheetnames:
        raise ValueError("'급여' 시트를 찾을 수 없습니다. 총인건비 파일이 맞는지 확인해주세요.")
    return wb


def list_worksites(wb):
    """급여 시트에 존재하는 사업장명 목록"""
    sal = wb["급여"]
    seen = []
    for r in range(4, sal.max_row + 1):
        v = sal.cell(r, 2).value
        if v and v not in seen:
            seen.append(v)
    return seen


def collect_payslip_rows(wb, worksite_names):
    """선택 사업장의 급여 + 기타비용 데이터를 개인별로 모은다."""
    from collections import defaultdict
    sal = wb["급여"]
    rows = []
    for r in range(4, sal.max_row + 1):
        site = sal.cell(r, 2).value
        if site not in worksite_names:
            continue
        name = sal.cell(r, 4).value
        if not name:
            continue
        rows.append({
            "site": site,
            "name": str(name).strip(),
            "pay": [sal.cell(r, c).value or 0 for c in range(5, 15)],  # E~N
        })

    extra = defaultdict(lambda: {"retire_annual": 0, "incentive": 0})
    unassigned = []
    if "기타비용" in wb.sheetnames:
        etc = wb["기타비용"]
        for r in range(2, etc.max_row + 1):
            site = etc.cell(r, 2).value
            if site not in worksite_names:
                continue
            nm = etc.cell(r, 4).value
            amt = etc.cell(r, 13).value or 0   # M: 계
            reason = etc.cell(r, 14).value or ""
            if not nm:
                if amt:
                    unassigned.append((str(reason).strip() or "사유없음", amt))
                continue
            nm = str(nm).strip()
            extra[nm]["retire_annual"] += etc.cell(r, 10).value or 0     # J: 퇴사자 연차수당
            extra[nm]["incentive"] += (
                (etc.cell(r, 7).value or 0)      # G: 교통비
                + (etc.cell(r, 8).value or 0)    # H: 고객사지급
                + (etc.cell(r, 9).value or 0)    # I: 기타
                + (etc.cell(r, 11).value or 0)   # K: 명절 교통비
            )

    rows.sort(key=lambda x: x["name"])
    for row in rows:
        e = extra.get(row["name"], {"retire_annual": 0, "incentive": 0})
        row["retire_annual"] = e["retire_annual"]
        row["incentive"] = e["incentive"]
    return rows, unassigned


def build_payslip_workbook(rows, title_text):
    """급여내역서 엑셀 생성 → BytesIO 반환"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    medium = Side(style="medium")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "급여"

    ws["B2"] = title_text
    ws["B2"].font = Font(name=PAYSLIP_FONT, size=16, bold=True)
    ws.row_dimensions[2].height = 25.2

    for i, label in enumerate(PAYSLIP_HEADERS):
        col = 2 + i
        cell = ws.cell(4, col, label)
        cell.font = Font(name=PAYSLIP_FONT, size=8)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        ws.cell(5, col).border = border_thin
        ws.cell(5, col).fill = header_fill
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.row_dimensions[4].height = 13.8
    ws.row_dimensions[5].height = 13.8

    start_row = 6
    for idx, row in enumerate(rows):
        r = start_row + idx
        ws.row_dimensions[r].height = 22.8
        values = (
            [row["site"], row["name"]]
            + list(row["pay"])
            + [None, row["retire_annual"] or None, row["incentive"] or None]
        )
        for i, v in enumerate(values):
            cell = ws.cell(r, 2 + i, v)
            cell.font = Font(name=PAYSLIP_FONT, size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_thin
            if i >= 2:
                cell.number_format = "###,##0"
        q = ws.cell(r, 17, f"=SUM(D{r}:P{r})")
        q.font = Font(name=PAYSLIP_FONT, size=8)
        q.alignment = Alignment(horizontal="center", vertical="center")
        q.border = border_thin
        q.number_format = "###,##0"

    last = start_row + len(rows) - 1
    total_row = last + 1
    ws.row_dimensions[total_row].height = 19.8
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    tc = ws.cell(total_row, 2, "합계")
    tc.font = Font(name=PAYSLIP_FONT, size=12, bold=True)
    tc.fill = header_fill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = Border(left=medium, right=thin, top=medium, bottom=medium)
    ws.cell(total_row, 3).border = Border(left=thin, right=medium, top=medium, bottom=medium)

    for col in range(4, 18):
        L = get_column_letter(col)
        c = ws.cell(total_row, col, f"=SUM({L}{start_row}:{L}{last})")
        c.font = Font(name=PAYSLIP_FONT, size=10 if col == 17 else 8, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "###,##0"
        c.border = Border(left=thin, right=thin, top=medium, bottom=medium)

    ws.column_dimensions["A"].width = 4.5
    ws.column_dimensions["B"].width = 18.0
    ws.column_dimensions["C"].width = 11.1
    for col in range(4, 17):
        ws.column_dimensions[get_column_letter(col)].width = 8.6
    ws.column_dimensions["Q"].width = 13.5
    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def read_salary_report(file):
    """급여자료보고서 핵심 데이터 구조적 추출"""
    try:
        engine = "xlrd" if str(file.name).endswith(".xls") else "openpyxl"
        df = pd.read_excel(file, engine=engine, header=None)

        def gv(row, col):
            try:
                v = df.iloc[row, col]
                return "" if str(v) == "nan" else v
            except:
                return ""

        def fmt(v):
            try:
                return f"{int(float(v)):,}원"
            except:
                return str(v)

        r = []
        r.append(f"급여월: {gv(3,17)} / 거래처: {gv(3,7)}")
        r.append("")
        r.append("[매출액]")
        r.append(f"  당월: {fmt(gv(6,8))} / 전월: {fmt(gv(6,10))} / 증감: {fmt(gv(6,13))}")
        if gv(6,15): r.append(f"  사유: {gv(6,15)}")
        r.append("")
        r.append("[상용직 급여 상세]")
        items = [(8,"기본급"),(9,"식대비(비과세)"),(10,"연장수당"),(11,"성과급"),(12,"상여금"),(13,"기타수당 및 급여")]
        for ri, name in items:
            ca, pa, da = gv(ri,8), gv(ri,10), gv(ri,13)
            cc, pc, dc = gv(ri,6), gv(ri,9), gv(ri,11)
            if ca or pa:
                r.append(f"  {name}: 당월 {fmt(ca)}({cc}명) / 전월 {fmt(pa)}({pc}명) / 증감 {fmt(da)}({dc}명)")
        ca2, pa2, da2 = gv(17,8), gv(17,10), gv(17,13)
        r.append(f"  → 상용직 계: 당월 {fmt(ca2)} / 전월 {fmt(pa2)} / 증감 {fmt(da2)}")
        r.append("")
        r.append("[일용직 급여]")
        r.append(f"  기본급: 당월 {fmt(gv(18,8))}({gv(18,6)}명) / 전월 {fmt(gv(18,10))}({gv(18,9)}명) / 증감 {fmt(gv(18,13))}({gv(18,11)}명)")
        r.append("")
        r.append("[급여지급율 - 가지급금 포함]")
        r.append(f"  당월: {gv(23,8)} / 전월: {gv(23,10)} / 변동: {gv(23,13)}")
        if gv(23,15): r.append(f"  사유: {gv(23,15)}")
        r.append("[급여지급율 - 가지급금 제외]")
        r.append(f"  당월: {gv(24,8)} / 전월: {gv(24,10)} / 변동: {gv(24,13)}")
        return "\n".join(r)
    except Exception as e:
        return f"읽기 오류: {str(e)}"

# ─────────────────────────────────────────
# 데이터 저장/로드 (GitHub 영구 저장)
# ─────────────────────────────────────────
DATA_FILE = "data.json"

# Streamlit Cloud는 앱이 잠들었다 깨어날 때 로컬 파일이 초기화되므로,
# GITHUB_TOKEN이 있으면 data.json을 GitHub 저장소에도 함께 저장해서
# 재시작해도 데이터가 유지되게 한다.
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = "mj910326/sabup"
GITHUB_PATH = "data.json"
GITHUB_BRANCH = "main"

def _github_headers():
    return {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}

def github_load():
    """GitHub 저장소에서 data.json을 읽어온다. 실패하면 None 반환."""
    if not GITHUB_TOKEN:
        return None
    try:
        import requests, base64
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_PATH}"
        resp = requests.get(url, headers=_github_headers(), params={"ref": GITHUB_BRANCH}, timeout=10)
        if resp.status_code == 200:
            content = resp.json()
            decoded = base64.b64decode(content["content"]).decode("utf-8")
            return json.loads(decoded)
        return None
    except Exception:
        return None

def github_save(data):
    """data.json을 GitHub 저장소에 커밋한다. 결과를 세션에 기록한다."""
    if not GITHUB_TOKEN:
        st.session_state["_gh_status"] = ("none", "GITHUB_TOKEN이 설정되지 않아 영구 저장이 꺼져 있습니다.")
        return False
    try:
        import requests, base64
        url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_PATH}"
        get_resp = requests.get(url, headers=_github_headers(), params={"ref": GITHUB_BRANCH}, timeout=15)
        sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None
        content_str = json.dumps(data, ensure_ascii=False, indent=2)
        b64 = base64.b64encode(content_str.encode("utf-8")).decode("utf-8")
        payload = {
            "message": f"data 자동 저장 {datetime.now().isoformat(timespec='seconds')}",
            "content": b64,
            "branch": GITHUB_BRANCH,
        }
        if sha:
            payload["sha"] = sha
        put_resp = requests.put(url, headers=_github_headers(), json=payload, timeout=15)
        if put_resp.status_code in (200, 201):
            st.session_state["_gh_status"] = ("ok", f"GitHub 저장 완료 ({datetime.now().strftime('%H:%M:%S')})")
            return True
        else:
            try:
                detail = put_resp.json().get("message", "")
            except Exception:
                detail = put_resp.text[:200]
            st.session_state["_gh_status"] = (
                "fail", f"GitHub 저장 실패 (코드 {put_resp.status_code}): {detail}"
            )
            return False
    except Exception as e:
        st.session_state["_gh_status"] = ("fail", f"GitHub 저장 오류: {type(e).__name__} - {str(e)[:200]}")
        return False

def show_save_status():
    """마지막 저장 결과를 화면에 표시한다."""
    status = st.session_state.get("_gh_status")
    if not status:
        return
    kind, msg = status
    if kind == "ok":
        st.success(f"💾 {msg}")
    elif kind == "fail":
        st.error(f"⚠️ {msg}\n\n데이터가 영구 저장되지 않았습니다. Secrets의 GITHUB_TOKEN을 확인해주세요.")
    else:
        st.warning(f"⚠️ {msg}")

def get_default_data():
    return {
        "settings": {
            "system_title": "사업장 관리 시스템",
            "sidebar_title": "사업장 관리 시스템",
            "font": "기본",
            "bg_color": "#ffffff",
            "accent_color": "#1f77b4"
        },
        "clients": {
            "풀무원": {
                "sites": {
                    "SR6": {"meetings": [], "issues": [], "history": []},
                    "동탄에듀": {"meetings": [], "issues": [], "history": []},
                    "넥슨": {"meetings": [], "issues": [], "history": []}
                }
            },
            "용마로지스": {"sites": {"용마로지스": {"meetings": [], "issues": [], "history": []}}},
            "미리디": {"sites": {"미리디": {"meetings": [], "issues": [], "history": []}}},
            "국세": {"sites": {"국세": {"meetings": [], "issues": [], "history": []}}},
            "한국환경공단": {"sites": {"한국환경공단": {"meetings": [], "issues": [], "history": []}}}
        },
        "checklist": {
            "items": ["근태", "청구서", "매출세금계산서 발행", "매출세금계산서 결재", "공문발송", "품의", "운영비 결재상신", "법인카드 전표", "매입세금계산서 발행", "매입세금계산서 결재", "급여작업", "급여내역서", "급여자료보고서 결재", "급여확정"],
            "site_enabled": {},
            "checked": {},
            "site_visible": {}
        }
    }

def _fill_defaults(loaded):
    default = get_default_data()
    for key in default:
        if key not in loaded:
            loaded[key] = default[key]
    if "site_visible" not in loaded.get("checklist", {}):
        loaded["checklist"]["site_visible"] = {}
    return loaded

def load_data():
    # 1순위: GitHub에 저장된 최신 데이터 (앱이 재시작돼도 유지됨)
    gh = github_load()
    if gh is not None:
        return _fill_defaults(gh)

    # 2순위: 로컬 파일 (같은 세션 내에서는 유효)
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                content_str = f.read().strip()
                if not content_str:
                    raise ValueError("빈 파일")
                loaded = json.loads(content_str)
                return _fill_defaults(loaded)
        except Exception:
            import shutil
            try:
                shutil.copy(DATA_FILE, DATA_FILE + ".bak")
            except:
                pass
            return get_default_data()
    return get_default_data()

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    github_save(data)

import os
API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
data = load_data()
settings = data["settings"]

FONTS = {
    "기본": "sans-serif",
    "나눔고딕": "'Nanum Gothic', sans-serif",
    "나눔명조": "'Nanum Myeongjo', serif",
    "고딕A1": "'Gothic A1', sans-serif",
    "IBM Plex": "'IBM Plex Sans KR', sans-serif",
}

font_css = FONTS.get(settings.get("font", "기본"), "sans-serif")
bg_color = settings.get("bg_color", "#ffffff")
accent = settings.get("accent_color", "#1f77b4")

st.set_page_config(page_title=settings["system_title"], page_icon="🏢", layout="wide")
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nanum+Gothic&family=Nanum+Myeongjo&family=Gothic+A1&family=IBM+Plex+Sans+KR&display=swap');
html, body, [class*="css"] {{ font-family: {font_css} !important; background-color: {bg_color} !important; }}
.card {{ background: white; border-left: 4px solid {accent}; padding: 12px 16px; border-radius: 8px; margin-bottom: 8px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown(f"## 🏢 {settings['sidebar_title']}")
    st.markdown("---")
    menu_items = ["🏠 홈"]
    menu_map = {}
    for client, cdata in data["clients"].items():
        menu_items.append(f"👥 {client}")
        menu_map[f"👥 {client}"] = (client, None)
        for site in cdata["sites"].keys():
            label = f"　🍽️ {site}"
            menu_items.append(label)
            menu_map[label] = (client, site)
    menu_items.append("📋 월초체크리스트")
    menu_items.append("⚙️ 설정")
    menu = st.radio("메뉴", menu_items)

# ─────────────────────────────────────────
# 🏠 홈
# ─────────────────────────────────────────
if menu == "🏠 홈":
    st.title(f"🏢 {settings['system_title']}")
    st.markdown("---")
    for client, cdata in data["clients"].items():
        st.markdown(f"### 👥 {client}")
        cols = st.columns(max(len(cdata["sites"]), 1))
        for i, (site, sdata) in enumerate(cdata["sites"].items()):
            with cols[i]:
                issue_count = len([x for x in sdata["issues"] if x.get("status") == "미처리"])
                st.markdown(f"<div class='card'><b>🍽️ {site}</b><br>📝 회의록: {len(sdata['meetings'])}건<br>⚠️ 미처리 이슈: {issue_count}건<br>📋 히스토리: {len(sdata['history'])}건</div>", unsafe_allow_html=True)
    st.markdown("---")
    st.subheader("⚠️ 전체 미처리 이슈")
    found = False
    for client, cdata in data["clients"].items():
        for site, sdata in cdata["sites"].items():
            for issue in sdata["issues"]:
                if issue.get("status") == "미처리":
                    found = True
                    st.error(f"**[{client} - {site}]** {issue['date']} - {issue['title']}")
    if not found:
        st.success("✅ 현재 미처리 이슈가 없습니다!")

# ─────────────────────────────────────────
# 👥 고객사
# ─────────────────────────────────────────
elif menu in menu_map and menu_map[menu][1] is None:
    client = menu_map[menu][0]
    cdata = data["clients"][client]
    st.title(f"👥 {client}")
    st.markdown("---")
    cols = st.columns(max(len(cdata["sites"]), 1))
    for i, (site, sdata) in enumerate(cdata["sites"].items()):
        with cols[i]:
            issue_count = len([x for x in sdata["issues"] if x.get("status") == "미처리"])
            st.markdown(f"<div class='card'><b>🍽️ {site}</b><br>📝 회의록: {len(sdata['meetings'])}건<br>⚠️ 미처리 이슈: {issue_count}건<br>📋 히스토리: {len(sdata['history'])}건</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────
# 🍽️ 사업장
# ─────────────────────────────────────────
elif menu in menu_map and menu_map[menu][1] is not None:
    client, site = menu_map[menu]
    sdata = data["clients"][client]["sites"][site]
    sk = f"{client}_{site}"

    st.title(f"🍽️ {client} - {site}")
    st.markdown("---")

    tab1, tab5, tab2, tab3, tab4 = st.tabs(
        ["📊 AI 분석", "📄 급여내역서", "📝 회의록", "⚠️ 이슈사항", "📋 히스토리"]
    )

    with tab1:
        st.markdown("### 📊 AI 분석")
        has_step1 = f"s1_{sk}" in st.session_state
        if has_step1:
            s1_month = st.session_state.get(f"s1_month_{sk}", "")
            st.success(f"✅ Step1 청구서 분석 결과 있음 {s1_month} → Step2/3에서 참고 가능")
            if st.button("🗑️ Step1 결과 초기화", key=f"clear1_{sk}"):
                del st.session_state[f"s1_{sk}"]
                st.rerun()

        step = st.radio("분석 단계", [
            "Step 1 - 청구서 분석",
            "Step 2 - 급여자료보고서 분석",
            "Step 3 - 매출이익 분석"
        ], horizontal=True, key=f"step_{sk}")

        # ── Step 1 ──
        if step == "Step 1 - 청구서 분석":
            st.info(f"📌 {site} 관련 데이터만 자동 필터링됩니다.")
            memo = st.text_area(
                "📝 참고사항 메모 (선택) - AI 분석 시 반영됩니다",
                placeholder="예: 전월 연차수당 과지급건 당월 환급 / 안전인센티브 지급 등",
                key=f"memo_{sk}", height=70
            )
            col1, col2 = st.columns(2)
            with col1:
                prev_file = st.file_uploader("전월 청구서", type=["xlsx", "xls"], key=f"p1_{sk}")
            with col2:
                curr_file = st.file_uploader("당월 청구서", type=["xlsx", "xls"], key=f"c1_{sk}")

            if prev_file and curr_file:
                prev_month = extract_month_from_filename(prev_file.name) or "전월"
                curr_month = extract_month_from_filename(curr_file.name) or "당월"
                st.success(f"✅ {prev_month} / {curr_month} 업로드 완료!")

                if st.button("🤖 청구서 분석 시작", type="primary", key=f"btn1_{sk}"):
                    with st.spinner("분석 중..."):
                        try:
                            prev_sheets = read_excel_sheets(prev_file)
                            curr_sheets = read_excel_sheets(curr_file)
                            prev_extra = read_extra_cost_sheet(prev_file, site)
                            curr_extra = read_extra_cost_sheet(curr_file, site)

                            def get_filtered_text(sheets, label):
                                text = ""
                                for sheet_name, df in sheets.items():
                                    if sheet_name == "기타비용":
                                        continue
                                    salary_df, sev_df = filter_site_data(df, site)
                                    if not salary_df.empty:
                                        try:
                                            s = salary_df.to_string()
                                        except:
                                            s = salary_df.to_string(encoding="utf-8") if hasattr(salary_df.to_string, "__call__") else str(salary_df)
                                        text += f"\n[{label} - {sheet_name} - 급여]\n{s}\n"
                                    if not sev_df.empty:
                                        try:
                                            s = sev_df.to_string()
                                        except:
                                            s = str(sev_df)
                                        text += f"\n[{label} - {sheet_name} - 퇴직금]\n{s}\n"
                                return text

                            prev_text = get_filtered_text(prev_sheets, prev_month)
                            curr_text = get_filtered_text(curr_sheets, curr_month)

                            if not prev_text and not curr_text:
                                st.warning("해당 사업장 데이터를 찾지 못했습니다.")
                            else:
                                client_obj = anthropic.Anthropic(api_key=API_KEY)
                                prompt = f"""당신은 노무/급여 전문가입니다. {site} 청구서(VAT 포함 기준)를 분석해주세요.

분석 기준월: 전월={prev_month}, 당월={curr_month}
모든 분석에서 "전월/당월" 대신 반드시 "{prev_month}/{curr_month}" 실제 월명을 사용하세요.

참고사항:
- 고정비 = 기본급 (시급직은 근무일수에 따라 변동)
- 변동비 = 연장/휴일 근무수당
- 퇴직금: 매달 6.5% 청구, 퇴사 시 법정 8.33%와 차액 익월 청구
  * 1년 이상 퇴사자 → 추가 청구 (퇴직금 증가)
  * 1년 미만 퇴사자 → 기청구분 환급 (퇴직금 감소)
- 기타비용: 이름 있는 항목 → 사유+금액 명시 / 이름 없는 항목 → 보험료 정산
- 인원변동: "중도퇴사 N명, 신규입사 N명" 후 실명 괄호 안에
- 악화/개선 절대 금지. 상승/하락/증가/감소만 사용

[담당자 메모]
{memo if memo else "없음"}

[{prev_month} 급여/퇴직금 데이터]
{prev_text}

[{curr_month} 급여/퇴직금 데이터]
{curr_text}

[{prev_month} 기타비용 상세]
{prev_extra}

[{curr_month} 기타비용 상세]
{curr_extra}

아래 형식으로 작성하세요. 각 항목은 줄을 나눠 간결하게:

## 📋 {site} 청구서 분석 ({curr_month} 기준, VAT 포함)

### 급여 현황
| 항목 | {prev_month} | {curr_month} | 증감 |
|------|-----:|-----:|-----:|
| 총 청구액(매출) | | | |
| 고정비(기본급) | | | |
| 변동비(연장/휴일) | | | |
| 일용직 | | | |
| 기타비용 | | | |

### 퇴직금 (별도)
| 항목 | {prev_month} | {curr_month} | 증감 |
|------|-----:|-----:|-----:|
| 퇴직금 청구액 | | | |

## 🔍 변동 원인
- **매출**: (원인)
- **고정비**: (인원현황 먼저, 실명 괄호)
- **변동비**: (연장/휴일 변동, 주요 인원 괄호)
- **기타비용**: (항목별 한 줄씩)
- **퇴직금**: (금액 패턴으로 원인 추론, 한 줄)
- **특이사항**: (없으면 생략)

숫자 천단위 콤마. 간결하게."""

                                msg = client_obj.messages.create(
                                    model="claude-sonnet-4-6",
                                    max_tokens=2000,
                                    messages=[{"role": "user", "content": prompt}]
                                )
                                st.session_state[f"s1_{sk}"] = msg.content[0].text
                                st.session_state[f"s1_month_{sk}"] = f"({prev_month}→{curr_month})"
                                st.session_state[f"s1_prev_month_{sk}"] = prev_month
                                st.session_state[f"s1_curr_month_{sk}"] = curr_month
                                st.rerun()

                        except Exception as e:
                            st.error(f"오류: {str(e)}")

            if f"s1_{sk}" in st.session_state:
                st.markdown("---")
                st.markdown(st.session_state[f"s1_{sk}"])
                if st.button("📋 히스토리 저장", key=f"sv1_{sk}"):
                    sdata["history"].append({
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "type": "청구서 분석",
                        "content": st.session_state[f"s1_{sk}"]
                    })
                    save_data(data)
                    show_save_status()

        # ── Step 2 ──
        elif step == "Step 2 - 급여자료보고서 분석":
            sal_file = st.file_uploader(
                "급여자료보고서 (선택 - 없어도 Step1 기반으로 분석 가능)",
                type=["xlsx", "xls"], key=f"sal_{sk}"
            )

            # 기준월 결정
            if sal_file:
                sal_month = extract_month_from_filename(sal_file.name) or "당월"
                match = re.search(r'(\d{4})년 (\d{2})월', sal_month)
                if match:
                    y, m = int(match.group(1)), int(match.group(2))
                    prev_m = 12 if m == 1 else m - 1
                    prev_y = y - 1 if m == 1 else y
                    sal_prev_month = f"{prev_y}년 {prev_m:02d}월"
                else:
                    sal_prev_month = "전월"
                st.success(f"✅ {sal_month} 급여자료보고서 업로드 완료!")
            else:
                sal_month = st.session_state.get(f"s1_curr_month_{sk}", "당월")
                sal_prev_month = st.session_state.get(f"s1_prev_month_{sk}", "전월")

            # 실행 가능 여부 안내
            if sal_file:
                st.info("📌 급여자료보고서 실제 수치 기준으로 분석합니다. (가장 정확)")
                can_run = True
            elif has_step1:
                st.info(
                    f"📌 파일 없이 **Step1 청구서 분석 결과({sal_prev_month} → {sal_month})** 기반으로 "
                    "증감사유를 생성합니다.\n\n"
                    "⚠️ 청구서는 VAT 포함, 급여자료보고서는 VAT 제외 기준이라 "
                    "**금액은 추정치**입니다. 증감사유 문구 초안 용도로 쓰시고, "
                    "정확한 금액이 필요하면 파일을 올려주세요."
                )
                can_run = True
            else:
                st.warning(
                    "⚠️ Step1 청구서 분석을 먼저 실행하거나, 급여자료보고서 파일을 올려주세요."
                )
                can_run = False

            if can_run:
                btn_label = "🤖 급여보고서 분석 시작" if sal_file else "🤖 Step1 기반으로 증감사유 생성"
                if st.button(btn_label, type="primary", key=f"btn2_{sk}"):
                    with st.spinner("분석 중..."):
                        try:
                            step1_ref = st.session_state.get(f"s1_{sk}", "없음 (청구서 미업로드)")
                            s1_prev = st.session_state.get(f"s1_prev_month_{sk}", sal_prev_month)
                            s1_curr = st.session_state.get(f"s1_curr_month_{sk}", sal_month)

                            if sal_file:
                                sal_text = read_salary_report(sal_file)
                                source_note = """※ 급여자료보고서 수치(VAT 제외) 기준으로 작성하세요.
※ 청구서는 참고용이며, 금액은 급여자료보고서 기준으로 작성하세요."""
                                data_block = f"""[급여자료보고서 수치]
{sal_text}

[청구서 분석 참고 - 인원/수당 상세 내용]
{step1_ref}"""
                                estimate_header = ""
                            else:
                                source_note = """※ 급여자료보고서 파일이 없습니다. Step1 청구서 분석 결과만으로 작성하세요.
※ 청구서는 VAT 포함 금액이므로, 급여자료보고서 기준(VAT 제외)으로 환산하세요.
   - VAT 제외 금액 = 청구서 금액 ÷ 1.1 (부가세 10% 제외)
   - 환산한 금액은 추정치이므로 표에 "(추정)" 표기
※ 급여지급율 등 청구서에서 알 수 없는 수치는 "파일 필요"로 표기하고 추측하지 마세요.
※ 인원/수당 변동 원인은 Step1 내용을 그대로 활용하세요."""
                                data_block = f"""[Step1 청구서 분석 결과 - 이것이 유일한 데이터입니다]
{step1_ref}"""
                                estimate_header = (
                                    "\n\n> ⚠️ **이 결과는 청구서(VAT 포함) 기반 추정치입니다.** "
                                    "정확한 금액은 급여자료보고서 파일을 올려 확인하세요.\n"
                                )

                            client_obj = anthropic.Anthropic(api_key=API_KEY)
                            prompt = f"""당신은 노무/급여 전문가입니다.
{site} {s1_curr} 급여자료보고서를 분석하여, 담당자가 ERP에 입력할 증감사유 멘트를 작성해주세요.

※ 이 멘트는 담당자가 ERP 급여자료보고서의 증감사유 칸에 그대로 입력할 내용입니다.
{source_note}

분석 기준월: 전월={s1_prev}, 당월={s1_curr}
모든 분석에서 반드시 "{s1_prev}/{s1_curr}" 실제 월명을 사용하세요.

작성 규칙:
- 증감사유 순서: 매출 → 기본급 → 식대 → 연장수당 → 기타수당 → 성과급/상여(있을때만) → 일용직
- 인원변동: "중도퇴사 N명, 신규입사 N명" 후 실명 괄호
- 기타수당 = 연장수당 제외한 모든 수당(휴일/심야/휴일연장 등)
- 변동 없는 항목 생략
- 악화/개선 금지. 상승/하락/증가/감소만 사용
- 각 항목 원인은 줄 나눠서 간결하게
- 데이터에 없는 수치는 절대 지어내지 말고 "-" 또는 "파일 필요"로 표기

{data_block}

아래 형식으로 작성하세요:

## 📝 {site} {s1_curr} 급여자료보고서

| 항목 | {s1_prev} | {s1_curr} | 증감 |
|------|-----:|-----:|-----:|
| 매출(VAT제외) | | | |
| 상용직 인원 | | | |
| 상용직 소득계 | | | |
| 일용직 인원 | | | |
| 일용직 소득계 | | | |
| 총 소득계 | | | |

## ✏️ ERP 입력용 증감사유

1. 매출: {s1_prev} N원 → {s1_curr} N원 (±N원)
   - (원인. 구체적으로)
2. 기본급: ±N원
   - (원인. 인원변동 상세, 실명 괄호)
3. 식대: ±N원 (변동 있을 때만)
   - (원인)
4. 연장수당: ±N원
   - (원인. 주요 인원 괄호)
5. 기타수당: ±N원 (있을 때만)
   - (원인: 휴일/심야수당 등 구체적으로)
6. 성과급/상여금: (있을 때만)
   - (원인)
7. 일용직: ±N원 (있을 때만)
   - (원인)

## 💡 급여지급율 분석 (★핵심★)

| 구분 | {s1_prev} | {s1_curr} | 변동 |
|------|-----:|-----:|-----:|
| 급여지급율(가지급금 포함) | % | % | %p |
| 급여지급율(가지급금 제외) | % | % | %p |

급여지급율이 변동한 이유:
- (인과관계로 설명. 예: "{s1_prev} 연차수당 과청구분 환급으로 {s1_curr} 매출 N원 감소, 노무비는 유지되어 지급율 N%p 상승")
- (추가 원인 있으면)

## 💬 상사 보고용 한줄 요약
> (핵심 1~2문장. 급여지급율 변동 원인 반드시 포함)

숫자 천단위 콤마. 간결하게."""

                            msg = client_obj.messages.create(
                                model="claude-sonnet-4-6",
                                max_tokens=2000,
                                messages=[{"role": "user", "content": prompt}]
                            )
                            st.session_state[f"s2_{sk}"] = estimate_header + msg.content[0].text

                        except Exception as e:
                            st.error(f"오류: {str(e)}")

            if f"s2_{sk}" in st.session_state:
                st.markdown("---")
                st.markdown(st.session_state[f"s2_{sk}"])
                if st.button("📋 히스토리 저장", key=f"sv2_{sk}"):
                    sdata["history"].append({
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "type": "급여보고서 분석",
                        "content": st.session_state[f"s2_{sk}"]
                    })
                    save_data(data)
                    show_save_status()

        # ── Step 3 ──
        elif step == "Step 3 - 매출이익 분석":
            if has_step1:
                st.info("📌 Step1 청구서 분석 결과를 참고합니다.")
            else:
                st.info("📌 Step1 없이 매출이익 파일만으로 분석합니다.")

            profit_file = st.file_uploader("거래처 월별 매출이익", type=["xlsx", "xls"], key=f"pf_{sk}")

            if profit_file:
                st.success("✅ 업로드 완료!")
                if st.button("🤖 매출이익 분석 시작", type="primary", key=f"btn3_{sk}"):
                    with st.spinner("분석 중..."):
                        try:
                            profit_text, prev_month, curr_month = read_profit_file(profit_file)
                            if not prev_month:
                                prev_month = st.session_state.get(f"s1_prev_month_{sk}", "전월")
                                curr_month = st.session_state.get(f"s1_curr_month_{sk}", "당월")
                            step1_ref = st.session_state.get(f"s1_{sk}", "없음 (청구서 미업로드)")

                            client_obj = anthropic.Anthropic(api_key=API_KEY)
                            prompt = f"""당신은 노무/급여 전문가입니다. {site} 매출이익(VAT 제외 기준)을 분석해주세요.

분석 기준월: 전월={prev_month}, 당월={curr_month}
모든 분석에서 "전월/당월" 대신 반드시 "{prev_month}/{curr_month}" 실제 월명을 사용하세요.

중요:
- 거래처 월별 매출이익 파일 수치(VAT 제외)를 기준으로 작성
- 청구서(VAT 포함)와 금액이 다를 수 있으므로 매출이익 파일 수치 우선
- 인원변동: "중도퇴사 N명, 신규입사 N명" 후 실명 괄호
- 퇴직충당금: 크게 증가→장기근속 퇴사자 차액분(8.33%-6.5%) 청구 추정 / 감소→단기근속 퇴사자 환급 또는 인원 감소 추정
- ERP 이슈사항 "가. 원인" 문구 절대 금지. 원인 내용만 바로 작성
- 악화/개선 절대 금지. 상승/하락/증가/감소만 사용
- 각 항목 줄 나눠서 간결하게

[거래처 월별 매출이익 (VAT 제외)]
{profit_text}

[Step1 청구서 분석 참고용]
{step1_ref}

아래 형식으로 작성하세요:

## 📊 {site} 매출이익 현황 ({curr_month})

| 항목 | {prev_month} | {curr_month} | 증감 |
|------|-----:|-----:|-----:|
| 인원(명) | | | |
| 매출(A) | | | |
| 직접원가(B) | | | |
| 직접노무비 | | | |
| 매출이익(A-B) | | | |
| 노무비율 | % | % | %p |
| 매출이익률 | % | % | %p |

## 💡 급여지급율 (노무비/매출액)
| 구분 | {prev_month} | {curr_month} | 변동 |
|------|-----:|-----:|-----:|
| 급여지급율 | % | % | %p |

- {prev_month}: 노무비 ÷ 매출액 = %
- {curr_month}: 노무비 ÷ 매출액 = %
- 변동 원인: (한 줄)

## 📝 ERP 입력용 - 당월 이슈사항

1. 인원: {prev_month} N명 → {curr_month} N명 (±N명)
   - (인원 변동 상세. 실명 괄호)

2. 매출: {prev_month} N원 → {curr_month} N원 (±N원)
   - (원인 1)
   - (원인 2, 있으면)

3. 매출이익: {prev_month} N원 → {curr_month} N원 (±N원)
   - (원인 1)
   - (원인 2, 있으면)

## 💬 상사 보고용 요약
> (핵심 1~2문장. 실제 월명 사용)

숫자 천단위 콤마. 비율 소수점 1자리. 간결하게."""

                            msg = client_obj.messages.create(
                                model="claude-sonnet-4-6",
                                max_tokens=2000,
                                messages=[{"role": "user", "content": prompt}]
                            )
                            st.session_state[f"s3_{sk}"] = msg.content[0].text

                        except Exception as e:
                            st.error(f"오류: {str(e)}")

            if f"s3_{sk}" in st.session_state:
                st.markdown("---")
                st.markdown(st.session_state[f"s3_{sk}"])
                if st.button("📋 히스토리 저장", key=f"sv3_{sk}"):
                    sdata["history"].append({
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "type": "매출이익 분석",
                        "content": st.session_state[f"s3_{sk}"]
                    })
                    save_data(data)
                    show_save_status()

            # ── 예상수지 ──
            if profit_file:
                st.markdown("---")
                st.markdown("### 📈 익월 예상수지")
                forecast_memo = st.text_area(
                    "📝 예상수지 참고 메모 (선택)",
                    placeholder="예: 6월 세척실계절수당 약 50만원 지급예정 / 신규입사 2명 예정 등",
                    key=f"fmemo_{sk}", height=70
                )
                if st.button("🔮 예상수지 분석", type="primary", key=f"btn_fc_{sk}"):
                    with st.spinner("예상수지 계산 중..."):
                        try:
                            fc_text, fc_prev, fc_curr = read_profit_file(profit_file)
                            fc_step1 = st.session_state.get(f"s1_{sk}", "없음")

                            client_fc = anthropic.Anthropic(api_key=API_KEY)
                            fc_prompt = f"""당신은 노무/급여 전문가입니다. {site}의 익월 예상수지를 산출해주세요.

데이터: {fc_text}

규칙:
- 최근 3~5개월 데이터 사용
- 상여금이 있는 달(명절 달)은 평균 계산에서 자동 제외
- 제외된 달이 있으면 어떤 달을 제외했는지 명시
- 담당자 메모 내용을 반영하여 예상수지 조정
- 모든 항목(인원/매출/급여/상여/퇴직충당금/잡급/노무비소계/복리후생비/접대비/법정복리후생비/인력수급비/지급수수료/소모품비/통신비/기타비용/경비소계/직접원가계/영업이익)을 채워주세요
- 숫자는 천단위 콤마
- 악화/개선 금지. 상승/하락/증가/감소만 사용

[담당자 메모 - 반드시 반영]
{forecast_memo if forecast_memo else "없음"}

[청구서 참고]
{fc_step1}

아래 형식으로 작성하세요:

## 📈 {site} 익월 예상수지

※ 평균 산출 기준: (사용한 월 명시, 제외한 달과 이유 명시)
※ 메모 반영 내용: (메모 있으면 어떻게 반영했는지)

| 항목 | 평균(N개월) | 익월 예상 | 비고 |
|------|-----:|-----:|------|
| 인원(명) | | | |
| 매출 | | | |
| 급여 | | | |
| 상여 | | 0 | 명절 없음 |
| 퇴직충당금 | | | |
| 잡급(일용직) | | | |
| 노무비 소계 | | | |
| 복리후생비 | | | |
| 접대비 | | | |
| 법정복리후생비 | | | |
| 인력수급비 | | | |
| 지급수수료 | | | |
| 소모품비 | | | |
| 통신비 | | | |
| 기타비용 | | | |
| 경비 소계 | | | |
| 직접원가 계 | | | |
| 영업이익(매출이익) | | | |
| 노무비율 | % | % | |
| 매출이익률 | % | % | |
| 급여지급율 | % | % | |

## 💬 예상수지 요약
> (핵심 내용 1~2문장. 전월 대비 변동 예상 포함)"""

                            fc_msg = client_fc.messages.create(
                                model="claude-sonnet-4-6",
                                max_tokens=2000,
                                messages=[{"role": "user", "content": fc_prompt}]
                            )
                            st.session_state[f"s3_fc_{sk}"] = fc_msg.content[0].text

                        except Exception as e:
                            st.error(f"오류: {str(e)}")

                if f"s3_fc_{sk}" in st.session_state:
                    st.markdown(st.session_state[f"s3_fc_{sk}"])
                    if st.button("📋 히스토리 저장 (예상수지)", key=f"sv_fc_{sk}"):
                        sdata["history"].append({
                            "date": datetime.now().strftime("%Y-%m-%d"),
                            "type": "예상수지",
                            "content": st.session_state[f"s3_fc_{sk}"]
                        })
                        save_data(data)
                        show_save_status()

    # ── 회의록 ──
    with tab2:
        st.markdown("### 📝 회의록")
        with st.expander("➕ 새 회의록 작성"):
            m_date = st.date_input("날짜", key=f"md_{sk}")
            m_title = st.text_input("제목", key=f"mt_{sk}")
            m_participants = st.text_input("참석자", key=f"mp_{sk}")
            m_content = st.text_area("내용", height=200, key=f"mc_{sk}")
            if st.button("저장", key=f"sm_{sk}"):
                if m_title:
                    sdata["meetings"].append({"date": str(m_date), "title": m_title, "participants": m_participants, "content": m_content})
                    save_data(data)
                    show_save_status()
                    st.rerun()
        for i, m in enumerate(reversed(sdata["meetings"])):
            with st.expander(f"📝 {m['date']} - {m['title']}"):
                st.write(f"**참석자:** {m.get('participants','')}")
                st.write(m["content"])
                if st.button("삭제", key=f"dm_{sk}_{i}"):
                    sdata["meetings"].pop(len(sdata["meetings"])-1-i)
                    save_data(data)
                    st.rerun()

    # ── 이슈사항 ──
    with tab3:
        st.markdown("### ⚠️ 이슈사항")
        with st.expander("➕ 새 이슈 등록"):
            i_date = st.date_input("날짜", key=f"id_{sk}")
            i_title = st.text_input("제목", key=f"it_{sk}")
            i_content = st.text_area("내용", height=150, key=f"ic_{sk}")
            i_status = st.selectbox("상태", ["미처리", "처리중", "완료"], key=f"is_{sk}")
            if st.button("등록", key=f"si_{sk}"):
                if i_title:
                    sdata["issues"].append({"date": str(i_date), "title": i_title, "content": i_content, "status": i_status})
                    save_data(data)
                    st.success("등록됨!")
                    st.rerun()
        filter_status = st.selectbox("상태 필터", ["전체", "미처리", "처리중", "완료"], key=f"if_{sk}")
        for i, issue in enumerate(reversed(sdata["issues"])):
            if filter_status != "전체" and issue["status"] != filter_status:
                continue
            color = {"미처리": "🔴", "처리중": "🟡", "완료": "🟢"}.get(issue["status"], "⚪")
            with st.expander(f"{color} {issue['date']} - {issue['title']} [{issue['status']}]"):
                st.write(issue["content"])
                new_status = st.selectbox("상태 변경", ["미처리", "처리중", "완료"],
                    index=["미처리", "처리중", "완료"].index(issue["status"]), key=f"uis_{sk}_{i}")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("상태 저장", key=f"uib_{sk}_{i}"):
                        sdata["issues"][len(sdata["issues"])-1-i]["status"] = new_status
                        save_data(data)
                        st.rerun()
                with c2:
                    if st.button("삭제", key=f"dib_{sk}_{i}"):
                        sdata["issues"].pop(len(sdata["issues"])-1-i)
                        save_data(data)
                        st.rerun()

    # ── 급여내역서 ──
    with tab5:
        st.markdown("### 📄 급여내역서 자동 생성")
        st.caption("총인건비 파일을 올리면 이 사업장 인원만 뽑아 급여내역서 엑셀로 만들어 드립니다.")

        tl_file = st.file_uploader(
            "총인건비 파일 (예: 급여작업중_2026_06_FNC_총인건비_○○○.xlsx)",
            type=["xlsx", "xlsm"], key=f"tl_{sk}"
        )

        if tl_file:
            try:
                tl_bytes = tl_file.read()
                tl_file.seek(0)
                wb_src = read_total_labor_file(tl_bytes)
                worksites = list_worksites(wb_src)

                # 이 사업장 키워드로 후보 자동 선택
                kws = SITE_KEYWORDS.get(site, {}).get("salary", [site])
                default_sel = [w for w in worksites if any(k in str(w) for k in kws)]

                st.success(f"✅ 총인건비 파일 읽기 완료 (사업장 {len(worksites)}개 감지)")

                sel_sites = st.multiselect(
                    "이 급여내역서에 포함할 사업장 (파일 안의 실제 사업장명)",
                    worksites,
                    default=default_sel,
                    key=f"tlsel_{sk}",
                    help="키워드로 자동 선택했습니다. 빠지거나 더해야 할 게 있으면 직접 조정하세요."
                )

                y, m = extract_ym_from_filename(tl_file.name)
                c1, c2 = st.columns([2, 1])
                with c1:
                    disp_name = st.text_input(
                        "제목에 넣을 사업장 표시명",
                        value=st.session_state.get(f"tlname_{sk}", site),
                        key=f"tlname_{sk}"
                    )
                with c2:
                    ym_default = f"{str(y)[2:]}년 {m}월" if y and m else ""
                    ym_text = st.text_input("연월", value=ym_default, key=f"tlym_{sk}",
                                            placeholder="26년 6월")

                title_text = f"▣ {disp_name} {ym_text} 급여내역서".replace("  ", " ").strip()
                st.caption(f"생성될 제목: **{title_text}**")

                if not sel_sites:
                    st.warning("사업장을 1개 이상 선택해주세요.")
                else:
                    rows, unassigned = collect_payslip_rows(wb_src, sel_sites)

                    if not rows:
                        st.error("선택한 사업장에 해당하는 인원이 없습니다.")
                    else:
                        # 미리보기
                        prev = []
                        for r in rows:
                            total = sum(r["pay"]) + r["retire_annual"] + r["incentive"]
                            prev.append({
                                "성명": r["name"],
                                "기본급": r["pay"][0],
                                "주휴수당": r["pay"][6],
                                "휴일수당": r["pay"][8],
                                "연장수당": r["pay"][9],
                                "퇴사자연차": r["retire_annual"],
                                "인센티브": r["incentive"],
                                "지급액계": total,
                            })
                        df_prev = pd.DataFrame(prev)

                        mc1, mc2 = st.columns(2)
                        mc1.metric("인원", f"{len(rows)}명")
                        mc2.metric("지급액계 합계", f"{int(df_prev['지급액계'].sum()):,}원")

                        st.dataframe(df_prev, use_container_width=True, hide_index=True)

                        if unassigned:
                            st.info(
                                "ℹ️ 기타비용 중 **개인 이름이 없는 항목**은 급여내역서에 들어가지 않습니다 "
                                "(청구서에는 별도 반영):\n\n"
                                + "\n".join(f"- {rsn}: {int(amt):,}원" for rsn, amt in unassigned)
                            )

                        st.warning(
                            "⚠️ **가지급 열은 비어 있습니다.** 원본 데이터에 없는 항목이라 "
                            "다운로드 후 직접 입력하셔야 합니다. "
                            "그 밖에 수기 조정할 부분도 엑셀에서 수정하세요."
                        )

                        buf = build_payslip_workbook(rows, title_text)
                        fname = f"{ym_text.replace(' ', '_')}_{disp_name}_급여내역서.xlsx".lstrip("_")
                        st.download_button(
                            "📥 급여내역서 엑셀 다운로드",
                            data=buf,
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            key=f"dl_{sk}"
                        )

            except Exception as e:
                st.error(f"파일 처리 오류: {type(e).__name__} - {str(e)}")
                st.caption("총인건비 파일에 '급여' 시트와 '기타비용' 시트가 있는지 확인해주세요.")
        else:
            with st.expander("ℹ️ 어떤 파일을 올려야 하나요?"):
                st.markdown("""
**총인건비 파일**을 올려주세요. 아래 시트가 있어야 합니다.

| 시트 | 사용하는 내용 |
|---|---|
| **급여** | 사업장·성명·기본급~연장수당 (개인별 급여 명세) |
| **기타비용** | 퇴사자 연차수당, 교통비·고객사지급(→인센티브) |

**자동 매핑 규칙**
- 기본급 ~ 연장수당 10개 항목 → 급여 시트에서 그대로
- 퇴사자연차수당 → 기타비용 '퇴사자 연차수당' 열
- 인센티브 → 기타비용의 교통비 + 고객사지급 + 기타 + 명절교통비 합산
- 지급액계 → 위 전체 합 (엑셀 수식으로 들어감)
- **가지급 → 빈칸** (원본에 없는 항목, 직접 입력 필요)
                """)

    # ── 히스토리 ──
    with tab4:
        st.markdown("### 📋 AI 분석 히스토리")
        if not sdata["history"]:
            st.info("저장된 분석 결과가 없습니다.")
        for i, h in enumerate(reversed(sdata["history"])):
            with st.expander(f"📄 {h['date']} - {h['type']}"):
                st.markdown(h["content"])
                if st.button("삭제", key=f"dh_{sk}_{i}"):
                    sdata["history"].pop(len(sdata["history"])-1-i)
                    save_data(data)
                    st.rerun()


# ─────────────────────────────────────────
# 📋 월초체크리스트
# ─────────────────────────────────────────
elif menu == "📋 월초체크리스트":
    st.title("📋 월초체크리스트")
    st.markdown("---")

    # 체크리스트 데이터 초기화
    if "checklist" not in data:
        data["checklist"] = {
            "items": ["근태", "청구서", "매출세금계산서 발행", "매출세금계산서 결재",
                     "공문발송", "품의", "운영비 결재상신", "법인카드 전표",
                     "매입세금계산서 발행", "매입세금계산서 결재", "급여작업",
                     "급여내역서", "급여자료보고서 결재", "급여확정"],
            "site_enabled": {},
            "checked": {},
            "site_visible": {}
        }

    cl = data["checklist"]
    cl.setdefault("site_visible", {})
    cl.setdefault("layout", {})
    _ldef = {
        "item_width": 2.0,
        "site_width": 1.0,
        "padding": 8,
        "align": "center",
        "spacer": 0.0,
        "row_height": 38,
        "font_size": 14,
        "header_bg": "#1f77b4",
        "header_text": "#ffffff",
        "item_bg": "#f0f2f6",
        "item_text": "#262730",
        "grid_color": "#c8ccd4",
        "disabled_bg": "#fafafa",
    }
    for _k, _v in _ldef.items():
        cl["layout"].setdefault(_k, _v)
    items = cl.get("items", [])

    # 전체 사업장 목록
    all_sites = []
    for c, cd in data["clients"].items():
        for s in cd["sites"].keys():
            all_sites.append(f"{c} - {s}")

    # site_enabled 초기화 (새 사업장/항목만 추가, 기존 설정 유지)
    for site_key in all_sites:
        if site_key not in cl["site_enabled"]:
            cl["site_enabled"][site_key] = {}
        for item in items:
            if item not in cl["site_enabled"][site_key]:
                cl["site_enabled"][site_key][item] = True

    # checked 초기화
    for site_key in all_sites:
        if site_key not in cl["checked"]:
            cl["checked"][site_key] = {}

    # site_visible 초기화 (새 사업장은 기본으로 표시)
    for site_key in all_sites:
        if site_key not in cl["site_visible"]:
            cl["site_visible"][site_key] = True

    # 실제로 화면에 표시할 사업장만 필터링
    visible_sites = [s for s in all_sites if cl["site_visible"].get(s, True)]

    # 현재 월
    curr_month = datetime.now().strftime("%Y년 %m월")

    # 상단 버튼
    col_title, col_btn = st.columns([3, 1])
    with col_title:
        st.subheader(f"📅 {curr_month}")
    with col_btn:
        if st.button("🔄 전체 초기화", type="secondary", key="cl_reset"):
            for site_key in all_sites:
                cl["checked"][site_key] = {}
            save_data(data)
            # 세션 상태 전체 체크박스 키 삭제
            for k in list(st.session_state.keys()):
                if k.startswith("cb_") or k.startswith("cl_"):
                    del st.session_state[k]
            st.success("✅ 전체 초기화 완료!")
            st.rerun()

    # ── 표시할 사업장 선택 ──
    with st.expander("🏢 표시할 사업장 선택 (체크 해제하면 표에서 빠집니다)"):
        with st.form("toggle_site_visible"):
            vis_cols = st.columns(4)
            vis_toggles = {}
            for i, site_key in enumerate(all_sites):
                with vis_cols[i % 4]:
                    site_short = site_key.split(" - ")[1] if " - " in site_key else site_key
                    vis_toggles[site_key] = st.checkbox(
                        site_short,
                        value=cl["site_visible"].get(site_key, True),
                        key=f"vis_{site_key}"
                    )
            if st.form_submit_button("적용"):
                for site_key, val in vis_toggles.items():
                    cl["site_visible"][site_key] = val
                save_data(data)
                st.success("적용됨!")
                st.rerun()

    # ── 표 모양 조절 ──
    with st.expander("📐 표 모양 조절 (크기·간격·색상·글씨)"):
        layout = cl["layout"]
        with st.form("layout_form"):
            st.markdown("**칸 너비**")
            c1, c2 = st.columns(2)
            with c1:
                item_width = st.slider("업무항목 칸 너비", 0.3, 6.0,
                                       float(layout.get("item_width", 2.0)), step=0.1)
            with c2:
                site_width = st.slider("사업장 칸 너비", 0.3, 6.0,
                                       float(layout.get("site_width", 1.0)), step=0.1)
            spacer = st.slider(
                "오른쪽 빈 여백 (클수록 표가 왼쪽으로 좁아짐)",
                0.0, 20.0, float(layout.get("spacer", 0.0)), step=0.5
            )

            st.markdown("---")
            st.markdown("**행 높이 / 글씨**")
            d1, d2, d3 = st.columns(3)
            with d1:
                row_height = st.slider("행 높이(px)", 20, 80, int(layout.get("row_height", 38)))
            with d2:
                font_size = st.slider("글씨 크기(px)", 9, 24, int(layout.get("font_size", 14)))
            with d3:
                align = st.selectbox("글자 정렬", ["가운데", "왼쪽"],
                                     index=0 if layout.get("align", "center") == "center" else 1)

            st.markdown("---")
            st.markdown("**색상**")
            cc1, cc2, cc3 = st.columns(3)
            with cc1:
                header_bg = st.color_picker("헤더 배경", layout.get("header_bg", "#1f77b4"))
                header_text = st.color_picker("헤더 글자", layout.get("header_text", "#ffffff"))
            with cc2:
                item_bg = st.color_picker("업무항목 배경", layout.get("item_bg", "#f0f2f6"))
                item_text = st.color_picker("업무항목 글자", layout.get("item_text", "#262730"))
            with cc3:
                grid_color = st.color_picker("표 선 색상", layout.get("grid_color", "#c8ccd4"))
                disabled_bg = st.color_picker("제외된 칸 배경", layout.get("disabled_bg", "#fafafa"))

            if st.form_submit_button("적용", type="primary"):
                cl["layout"] = {
                    "item_width": item_width,
                    "site_width": site_width,
                    "spacer": spacer,
                    "row_height": row_height,
                    "font_size": font_size,
                    "align": "center" if align == "가운데" else "left",
                    "header_bg": header_bg,
                    "header_text": header_text,
                    "item_bg": item_bg,
                    "item_text": item_text,
                    "grid_color": grid_color,
                    "disabled_bg": disabled_bg,
                }
                save_data(data)
                st.success("적용됨!")
                st.rerun()

    st.markdown("---")

    layout = cl["layout"]
    align = layout.get("align", "center")
    spacer = float(layout.get("spacer", 0.0))
    row_h = int(layout.get("row_height", 38))
    font_size = int(layout.get("font_size", 14))
    header_bg = layout.get("header_bg", "#1f77b4")
    header_text = layout.get("header_text", "#ffffff")
    item_bg = layout.get("item_bg", "#f0f2f6")
    item_text = layout.get("item_text", "#262730")
    grid_color = layout.get("grid_color", "#c8ccd4")
    disabled_bg = layout.get("disabled_bg", "#fafafa")

    flex_align = "center" if align == "center" else "flex-start"
    text_align = "center" if align == "center" else "left"

    if not visible_sites:
        st.warning("표시할 사업장이 없습니다. 위에서 사업장을 선택해주세요.")
    else:
        # 표 영역만 스타일이 적용되도록 전용 컨테이너 사용
        try:
            grid = st.container(key="checklist_grid")
            SCOPE = ".st-key-checklist_grid "
        except TypeError:
            # 구버전 Streamlit은 key를 지원하지 않으므로 범위 지정 없이 적용
            grid = st.container()
            SCOPE = ""

        # 표 영역 전용 CSS - 모든 칸이 같은 높이/테두리를 갖는 진짜 표 형태
        st.markdown(f"""
        <style>
        {SCOPE}div[data-testid="stHorizontalBlock"] {{
            gap: 0 !important;
        }}
        {SCOPE}div[data-testid="stVerticalBlock"] {{
            gap: 0 !important;
        }}
        /* 각 칸 = 컬럼 자체에 테두리를 줘서 표처럼 선이 이어지게 */
        {SCOPE}div[data-testid="column"] {{
            border-right: 1px solid {grid_color};
            border-bottom: 1px solid {grid_color};
            height: {row_h}px;
            min-height: {row_h}px;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            overflow: hidden;
            padding: 0 !important;
        }}
        /* 맨 왼쪽 칸에만 왼쪽 선 */
        {SCOPE}div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:first-child {{
            border-left: 1px solid {grid_color};
        }}
        /* 첫 행(헤더)에만 윗선 */
        {SCOPE}div[data-testid="stHorizontalBlock"]:first-of-type > div[data-testid="column"] {{
            border-top: 1px solid {grid_color};
        }}
        /* 칸 안 요소들의 기본 여백 제거 */
        {SCOPE}div[data-testid="stElementContainer"],
        {SCOPE}div[data-testid="stVerticalBlock"] > div {{
            margin: 0 !important;
            padding: 0 !important;
            width: 100%;
        }}
        /* 체크박스를 칸 정중앙에 */
        {SCOPE}div[data-testid="stCheckbox"],
        {SCOPE}.stCheckbox {{
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 100% !important;
            height: {row_h}px !important;
            margin: 0 !important;
            padding: 0 !important;
        }}
        {SCOPE}div[data-testid="stCheckbox"] label,
        {SCOPE}.stCheckbox label {{
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 100% !important;
            height: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
        }}
        /* 셀 텍스트 */
        .cellbox {{
            width: 100%;
            height: {row_h}px;
            display: flex;
            align-items: center;
            justify-content: {flex_align};
            padding: 0 8px;
            box-sizing: border-box;
            font-size: {font_size}px;
            line-height: 1.15;
            text-align: {text_align};
            overflow: hidden;
        }}
        </style>
        """, unsafe_allow_html=True)

        def cellbox(text, bg, color, bold=False):
            weight = "700" if bold else "400"
            return (f"<div class='cellbox' style='background:{bg};color:{color};"
                    f"font-weight:{weight};'>{text}</div>")

        col_widths = [layout.get("item_width", 2.0)] + [layout.get("site_width", 1.0)] * len(visible_sites)
        if spacer > 0:
            col_widths = col_widths + [spacer]

        with grid:
            # ── 헤더 행 ──
            header_cols = st.columns(col_widths, gap="small")
            with header_cols[0]:
                st.markdown(cellbox("업무항목", header_bg, header_text, bold=True),
                            unsafe_allow_html=True)
            for i, site_key in enumerate(visible_sites):
                with header_cols[i+1]:
                    site_short = site_key.split(" - ")[1] if " - " in site_key else site_key
                    st.markdown(cellbox(site_short, header_bg, header_text, bold=True),
                                unsafe_allow_html=True)
            if spacer > 0:
                with header_cols[-1]:
                    st.markdown("<div class='cellbox'></div>", unsafe_allow_html=True)

            # ── 항목별 행 ──
            for idx, item in enumerate(items):
                row_cols = st.columns(col_widths, gap="small")
                with row_cols[0]:
                    st.markdown(cellbox(item, item_bg, item_text), unsafe_allow_html=True)

                for i, site_key in enumerate(visible_sites):
                    with row_cols[i+1]:
                        enabled = cl["site_enabled"].get(site_key, {}).get(item, True)
                        if enabled:
                            checked = cl["checked"].get(site_key, {}).get(item, False)
                            cb_key = f"cb_{site_key}_{item}"
                            if cb_key not in st.session_state:
                                st.session_state[cb_key] = checked
                            new_val = st.checkbox("", key=cb_key, label_visibility="collapsed")
                            if new_val != checked:
                                if site_key not in cl["checked"]:
                                    cl["checked"][site_key] = {}
                                cl["checked"][site_key][item] = new_val
                                save_data(data)
                        else:
                            st.markdown(cellbox("—", disabled_bg, "#9aa0a6"),
                                        unsafe_allow_html=True)

                if spacer > 0:
                    with row_cols[-1]:
                        st.markdown("<div class='cellbox'></div>", unsafe_allow_html=True)

    st.markdown("---")

    # ── 항목 관리 ──
    with st.expander("⚙️ 항목 관리"):
        st.markdown("**항목 추가**")
        with st.form("add_item"):
            new_item = st.text_input("새 항목 이름")
            if st.form_submit_button("추가"):
                if new_item and new_item not in cl["items"]:
                    cl["items"].append(new_item)
                    for sk in all_sites:
                        if sk not in cl["site_enabled"]:
                            cl["site_enabled"][sk] = {}
                        cl["site_enabled"][sk][new_item] = True
                    save_data(data)
                    st.success(f"'{new_item}' 추가됨!")
                    st.rerun()

        st.markdown("**항목 삭제**")
        with st.form("del_item"):
            del_item = st.selectbox("삭제할 항목", ["선택안함"] + items)
            if st.form_submit_button("삭제"):
                if del_item != "선택안함":
                    cl["items"].remove(del_item)
                    for sk in all_sites:
                        cl["site_enabled"].get(sk, {}).pop(del_item, None)
                        cl["checked"].get(sk, {}).pop(del_item, None)
                    save_data(data)
                    st.success(f"'{del_item}' 삭제됨!")
                    st.rerun()

        st.markdown("**사업장별 항목 활성화/비활성화**")
        st.caption("체크 해제하면 해당 사업장에서 항목이 '—'로 표시됩니다")
        sel_site = st.selectbox("사업장 선택", all_sites, key="cl_site_sel")
        if sel_site:
            with st.form("toggle_items"):
                toggles = {}
                for item in items:
                    cur = cl["site_enabled"].get(sel_site, {}).get(item, True)
                    toggles[item] = st.checkbox(item, value=cur, key=f"tog_{sel_site}_{item}")
                if st.form_submit_button("저장"):
                    if sel_site not in cl["site_enabled"]:
                        cl["site_enabled"][sel_site] = {}
                    for item, val in toggles.items():
                        cl["site_enabled"][sel_site][item] = val
                    save_data(data)
                    show_save_status()
                    st.rerun()

# ─────────────────────────────────────────
# ⚙️ 설정
# ─────────────────────────────────────────
elif menu == "⚙️ 설정":
    st.title("⚙️ 설정")
    st.markdown("---")
    tab_s1, tab_s2, tab_s3, tab_s4 = st.tabs(["🎨 화면 설정", "✏️ 이름 변경", "➕ 추가/삭제", "💾 저장 상태"])

    with tab_s4:
        st.markdown("### 💾 데이터 영구 저장 상태")
        st.caption("Streamlit Cloud는 앱이 재시작되면 로컬 파일이 초기화됩니다. "
                   "GitHub에 저장해야 다음날에도 데이터가 유지됩니다.")

        if not GITHUB_TOKEN:
            st.error(
                "⚠️ **GITHUB_TOKEN이 설정되지 않았습니다.**\n\n"
                "이 상태에서는 히스토리·체크리스트가 앱 재시작 시 사라집니다.\n\n"
                "Streamlit Cloud → 앱 Settings → Secrets 에 아래를 추가하세요:\n"
                "```\nGITHUB_TOKEN = \"ghp_...\"\n```"
            )
        else:
            st.info(f"토큰 설정됨 (끝 4자리: ...{GITHUB_TOKEN[-4:]}) / 저장소: {GITHUB_REPO}")

        if st.button("🔍 저장 연결 테스트", type="primary"):
            with st.spinner("GitHub 연결 확인 중..."):
                ok = github_save(data)
            show_save_status()
            if ok:
                st.balloons()
                st.markdown(
                    f"✅ 정상입니다. [저장소에서 data.json 확인하기]"
                    f"(https://github.com/{GITHUB_REPO}/blob/{GITHUB_BRANCH}/{GITHUB_PATH})"
                )
            else:
                st.markdown("""
**확인할 것:**
1. 토큰에 **repo** 권한이 있는지 (Fine-grained 토큰이면 Contents: Read and write)
2. 토큰이 만료되지 않았는지
3. 저장소 이름이 맞는지 (`mj910326/sabup`)
4. 기본 브랜치가 `main`인지 (`master`면 코드 수정 필요)
                """)

        st.markdown("---")
        st.markdown("**현재 저장된 데이터 요약**")
        total_hist = sum(
            len(sd.get("history", []))
            for cd in data["clients"].values()
            for sd in cd["sites"].values()
        )
        total_meet = sum(
            len(sd.get("meetings", []))
            for cd in data["clients"].values()
            for sd in cd["sites"].values()
        )
        total_issue = sum(
            len(sd.get("issues", []))
            for cd in data["clients"].values()
            for sd in cd["sites"].values()
        )
        m1, m2, m3 = st.columns(3)
        m1.metric("AI 분석 히스토리", f"{total_hist}건")
        m2.metric("회의록", f"{total_meet}건")
        m3.metric("이슈사항", f"{total_issue}건")

    with tab_s1:
        with st.form("display_form"):
            new_sys = st.text_input("시스템 제목", value=settings["system_title"])
            new_side = st.text_input("사이드바 제목", value=settings["sidebar_title"])
            new_font = st.selectbox("글씨체", list(FONTS.keys()), index=list(FONTS.keys()).index(settings.get("font", "기본")))
            c1, c2 = st.columns(2)
            with c1:
                new_bg = st.color_picker("배경색", value=settings.get("bg_color", "#ffffff"))
            with c2:
                new_accent = st.color_picker("강조색", value=settings.get("accent_color", "#1f77b4"))
            if st.form_submit_button("저장", type="primary"):
                data["settings"].update({"system_title": new_sys, "sidebar_title": new_side, "font": new_font, "bg_color": new_bg, "accent_color": new_accent})
                save_data(data)
                st.success("✅ 저장됨! 새로고침하세요.")

    with tab_s2:
        st.markdown("### 고객사 이름 변경")
        for client in list(data["clients"].keys()):
            with st.form(f"rc_{client}"):
                new_name = st.text_input(f"👥 {client}", value=client)
                if st.form_submit_button("변경"):
                    if new_name and new_name != client:
                        data["clients"][new_name] = data["clients"].pop(client)
                        save_data(data)
                        st.success("변경됨! 새로고침하세요.")
        st.markdown("### 사업장 이름 변경")
        for client, cdata in data["clients"].items():
            for site in list(cdata["sites"].keys()):
                with st.form(f"rs_{client}_{site}"):
                    new_name = st.text_input(f"🍽️ {client} > {site}", value=site)
                    if st.form_submit_button("변경"):
                        if new_name and new_name != site:
                            cdata["sites"][new_name] = cdata["sites"].pop(site)
                            save_data(data)
                            st.success("변경됨! 새로고침하세요.")

    with tab_s3:
        st.markdown("### ➕ 고객사 추가")
        with st.form("ac"):
            nc = st.text_input("새 고객사 이름")
            if st.form_submit_button("추가"):
                if nc and nc not in data["clients"]:
                    data["clients"][nc] = {"sites": {}}
                    save_data(data)
                    st.success(f"'{nc}' 추가됨! 새로고침하세요.")
        st.markdown("### ➕ 사업장 추가")
        with st.form("as"):
            tc = st.selectbox("고객사", list(data["clients"].keys()))
            ns = st.text_input("새 사업장 이름")
            if st.form_submit_button("추가"):
                if ns and ns not in data["clients"][tc]["sites"]:
                    data["clients"][tc]["sites"][ns] = {"meetings": [], "issues": [], "history": []}
                    save_data(data)
                    st.success(f"'{ns}' 추가됨! 새로고침하세요.")
        st.markdown("### 🗑️ 삭제")
        with st.form("del"):
            dc = st.selectbox("삭제할 고객사", ["선택안함"] + list(data["clients"].keys()))
            all_sites = [(c, s) for c, cd in data["clients"].items() for s in cd["sites"].keys()]
            ds = st.selectbox("삭제할 사업장", ["선택안함"] + [f"{c} - {s}" for c, s in all_sites])
            if st.form_submit_button("삭제", type="primary"):
                if dc != "선택안함":
                    del data["clients"][dc]
                    save_data(data)
                    st.success("삭제됨! 새로고침하세요.")
                elif ds != "선택안함":
                    c, s = ds.split(" - ", 1)
                    del data["clients"][c]["sites"][s]
                    save_data(data)
                    st.success("삭제됨! 새로고침하세요.")
